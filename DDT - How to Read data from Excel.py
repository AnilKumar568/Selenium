import openpyxl


# Reading Data From Excel file

path = r'C:\Users\Akumar4\Desktop\Test.xlsx'

workbook = openpyxl.load_workbook(path)
# sheet = workbook.active # capture active sheet details
sheet = workbook.get_sheet_by_name('Sheet1')

# Count No. of Rows & Cols
rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r, column=c).value, end='  ') # var name should be like row & column
    print()




